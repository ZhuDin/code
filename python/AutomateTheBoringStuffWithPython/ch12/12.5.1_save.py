import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames())
sheet = wb.sheetnames('Sheet')
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames())